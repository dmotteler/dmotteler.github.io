#!C:/Users/Del/AppData/Local/Programs/Python/Python36/python.exe
'''
    $Id: xlstojson.py,v 1.1 2022/05/13 19:26:19 dfm Exp dfm $
    $Log: xlstojson.py,v $
    Revision 1.1  2022/05/13 19:26:19  dfm
    Initial revision


    provide json of all tables in an xlsx file, given path to .xlsx
    if not path given, use latest EpiCurve_Count_Cases_Hospitalizations_Deaths2022
    in Downloads

    counties, tables, start date, and end date will be used to filter the data if present.

'''

import os, sys
from openpyxl import load_workbook
from datetime import datetime, timezone, timedelta
import json
import cgi

def get_table(wb, tabname, counties, start, end):
    try:
        ws = wb[tabname]
    except KeyError as e:
        return "No such worksheet"

    # opening in read-only mode gave just cell A1 without this reset:
    ws.reset_dimensions()

    tabhdrs = list(list(ws.iter_rows(max_row=1, values_only=True))[0])
    hdrdict = {x[1]:x[0] for x in enumerate(tabhdrs)}
    dtmcol = hdrdict.get('Date/Time Updated', None)
    if dtmcol:
        tabhdrs.pop(dtmcol)

    hdrnick = {}
    hdrnick['Cases'] = {
            'Earliest Specimen Collection Date':'Date',
            'Total Cases':'Cases',
            'Confirmed Cases':'Confirmed',
            'Probable Cases':'Probable',
            'Total Cases (7-Day Average)':'7-Day avg',
            '7-Day Case Count':'7-Day Count',
            '7-Day Case Rate':'7-Day Rate',
            '14-Day Case Count':'14-Day Count',
            '14-Day Case Rate':'14-Day Rate',
            }

    hdrnick['Hospitalizations'] = {
            'Hospital Admission Date':'Admission',
            'Hospitalizations':'Count',
            'Hospitalizations (7-Day Average)':'7-Day Avg',
            '7-Day Hospitalization Count':'7-Day Count',
            '7-Day Hospitalization Rate':'7-Day Rate',
            }

    hdrnick['Deaths'] = {
            'Earliest Specimen Collection Date':'Date',
            'Deaths':'Count',
            'Deaths (7-Day Average)':'7-Day Avg',
            '7-Day Death Count':'7-Day Count',
            '7-Day Death Rate':'7-Day Rate',
            }
    
    hdrnick['Data Dictionary'] = {
            'Variable Name': 'Variable',
            'Definition': 'Definition'
            }

    countycol = hdrdict.get('County', None)
    if countycol:
        tabhdrs.pop(countycol)

    hdrs = [hdrnick[tabname][x] for x in tabhdrs]

    doall = counties[0] == "all"
    if not doall:
        chkcounties = []
        for county in counties:
            if county == "Statewide":
                cntyname = county
            else:
                cntyname = "{} County".format(county)
            chkcounties.append(cntyname)

    clen = -len(" County")

    tabdata = {}

    nrows = 0
    dtms = {}
    ws.calculate_dimension(force=True)
    rownum = ws.max_row

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        # row[0] is date if there's a dtmcol - see if date in range
        if dtmcol:
            if start and row[0] < start:
                continue
            if end and end != "" and row[0] > end:
                continue

        # if we're not doing all counties, and this table has a county column,
        # and this row's county column is not in the ones requested, skip it.
        if (not doall) and countycol and row[countycol] not in chkcounties:
            continue

        if dtmcol:
            dtm = row.pop(dtmcol)
            if isinstance(dtm, datetime):
                dtm = "{:%b %d, %Y at %H:%M:%S}".format(dtm)
                if dtm not in dtms:
                    dtms[dtm] = 0
                dtms[dtm] += 1

        if countycol:
            county = row.pop(countycol)
            if county != "Statewide":
                county = county[:clen]

            if county not in tabdata:
                tabdata[county] = {}
            if tabname not in tabdata[county]:
                tabdata[county][tabname] = []
                tabdata[county][tabname].append(hdrs)
        
        rowout = []
        for cell in row:
            if isinstance(cell, datetime):
                cell = "{:%Y-%m-%d}".format(cell)
            rowout.append(cell)

        if countycol:
            tabdata[county][tabname].append(rowout)
        else:
            if tabname not in tabdata:
                tabdata[tabname] = []
            tabdata[tabname].append(rowout)
        nrows += 1

    if len(dtms) > 0:
        dtmtime, nocc = list(dtms.items())[0]
    else:
        dtmtime = None

    return dtmtime, tabdata

def main():
    xlfile = ''
    counties = []
    tables = []
    start = datetime(2022, 1, 1)
    end = datetime.today()

    form = cgi.FieldStorage()

    for kf in form:
        if kf.endswith("[]"):
            k = kf[:-2]
        else:
            k = kf

        if k == 'path':
            xlfile = form.getlist(kf)[0]

        elif k == 'county':
            for c in form.getlist(kf):
                counties.append(c)

        elif k == 'start':
            start = datetime.strptime(form.getlist(kf)[0], "%Y-%m-%d")

        elif k == 'end':
            end = datetime.strptime(form.getlist(kf)[0], "%Y-%m-%d")

        elif k.startswith('table'):
            for t in form.getlist(kf):
                tables.append(t)

    if xlfile == '':
        dldir = "C:/Users/Del/Downloads"
        _, _, files = next(os.walk(dldir))

        # reports = [x for x in files if x.startswith("WA_COVID19") and x.endswith(".xlsx")]
        reports = [x for x in files if x.startswith("EpiCurve_Count_Cases_Hospitalizations_Deaths") and x.endswith(".xlsx")]
        tim,fil = max([(os.path.getmtime("{}/{}".format(dldir, x)), x) for x in reports])

        mtim = datetime.fromtimestamp(tim, tz=timezone.utc)
        xlfile = "{}/{}".format(dldir, fil)

    else:
        mtim = datetime.fromtimestamp(os.path.getmtime(xlfile), tz=timezone.utc)
    
    xdir, xfn = os.path.split(xlfile)
    xfn, _ = os.path.splitext(xfn) 

    logfn = "C:/cygwin64/var/log/getxlsx.txt"
    log = open(logfn, "a")

    wb = load_workbook(xlfile, read_only=True)

    # default counties to all
    if len(counties) == 0:
        counties = ['all']

    # default tables to all
    if len(tables) < 1:
        tables = wb.sheetnames

    log.write("xlfile: {}, modtime {:%a, %d %b %Y %H:%M:%S %Z}\n".format(xlfile, mtim))
    log.write("start: {:%Y-%m-%d}, end: {:%Y-%m-%d}, counties: {}\n".format(start, end, ",".join(counties)))

    rv = {}
    rv['data'] = {}
    rv['meta'] = {}
    rv['meta']['xlfile'] = xlfile
    # the filetime returned is formatted for If-Modified-Since: Wed, 21 Oct 2015 07:28:00 GMT
    rv['meta']['filetime'] = "{:%a, %d %b %Y %H:%M:%S %Z}".format(mtim)
    rv['meta']['start'] = start.strftime("%Y-%m-%d")
    rv['meta']['end'] = end.strftime("%Y-%m-%d")

    rvcounties = []
    dtm = None
    for tablename in tables:
        if tablename == 'Data Dictionary':
            _, table = get_table(wb, tablename, counties, start, end)
            rv['meta'][tablename] = table[tablename]
        else:
            dtm, table = get_table(wb, tablename, counties, start, end)
            for county in table:
                if not county in rv['data']:
                    rv['data'][county] = {}
                if tablename not in rv['data'][county]:
                    rv['data'][county][tablename] = table[county][tablename][:]
                if not county in rvcounties:
                    rvcounties.append(county)

    rv['meta']['dtm'] = dtm
    rv['meta']['tables'] = ",".join(tables)
    rv['meta']['counties'] = ",".join(rvcounties)

    v = json.dumps(rv)
    log.write("rv length is {}\n".format(len(v)))

    jsonfile = "Epi.json"
    with open(jsonfile, "w") as jo:
        jo.write(v)

    print("Created {}".format(jsonfile))

if __name__ == "__main__":
    main()
